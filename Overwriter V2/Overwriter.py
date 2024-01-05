# -*- coding: utf-8 -*-
"""
Created on Thu Aug 31 23:18:09 2023

@author: Pariz Nanap Pandu Gumelar Pratama
"""
import openpyxl
import datetime
import os
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from openpyxl.styles import PatternFill
root = tk.Tk()
root.title('Tkinter Open File Dialog')
root.resizable(False, False)
root.geometry('400x300')
filename_source=""
filename_target=""
def select_source():
    global filename_source
    filename_source = fd.askopenfilename(
        title='Open a file',
        initialdir='/')

    showinfo(
        title='Selected File',
        message=filename_source
    )
    source=Label(root,text="Source Location:"+filename_source).place(x=20,y=90)
def select_target():
    global filename_target
    filename_target = fd.askopenfilename(
        title='Open a file',
        initialdir='/')

    showinfo(
        title='Selected File',
        message=filename_target
    )
    source=Label(root,text="Target Location:"+filename_target).place(x=20,y=150)
def run_program():
    try:
        f=filename_source
        if os.path.isfile(f):
            wrkbk = openpyxl.load_workbook(f)      
            sh = wrkbk.active
            maxprogress=sh.max_row-2
            ColNames = {}
            Current  = 0
            for COL in sh.iter_cols(1, sh.max_column):
                ColNames[COL[1].value] = Current
                Current += 1
            Keys = {}
            Currents  = 0
            for COL in sh.iter_cols(1, sh.max_column):
                Keys[COL[0].value] = Currents
                Currents += 1
            updatecol=[]
            for urow in range(sh.max_column):
                checkkey=sh.cell(row=1,column=urow+1).value
                if checkkey=="overwrite":
                    updatecol.append(urow+1)
            updatefill=[]
            for urow2 in range(sh.max_column):
                checkkey=sh.cell(row=1,column=urow2+1).value
                if checkkey=="update":
                    updatefill.append(urow2+1)
            print(updatefill)
            r=3
            iterations=0
            countfound=0
            for row in sh.iter_rows(min_row=3, min_col=1, max_row=sh.max_row, max_col=sh.max_column):
                found=0
                key=sh.cell(row=r,column=Keys['key']+1).value
                columnkey=sh.cell(row=2,column=Keys['key']+1).value
                g =filename_target
                if os.path.isfile(g):
                    wrkbk2 = openpyxl.load_workbook(g)
                    sh2 = wrkbk2.active
                    Colnames2={}
                    Current  = 0
                    for COL in sh2.iter_cols(1, sh2.max_column):
                        Colnames2[COL[0].value] = Current
                        Current += 1
                    if columnkey in Colnames2:
                        s=2
                        for row2 in sh2.iter_rows(min_row=2, min_col=1, max_row=sh2.max_row, max_col=sh2.max_column):
                            key2=sh2.cell(row=s,column=Colnames2[columnkey]+1).value
                            if key==key2:
                                for checky in range(len(updatecol)):
                                    checkcolumnsrc=sh.cell(row=2,column=updatecol[checky]).value
                                    if checkcolumnsrc in Colnames2:
                                        found=1
                                        sh2.cell(row=s,column=Colnames2[columnkey]+1).fill=PatternFill(start_color="CEFFC7", end_color="CEFFC7", fill_type = "solid")
                                        sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).value=sh.cell(row=r,column=updatecol[checky]).value
                                        sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).fill=PatternFill(start_color="CEFFC7", end_color="CEFFC7", fill_type = "solid")
                                        if isinstance(sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).value, datetime.datetime) is True:
                                            sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).number_format='yyyy-mm-dd'
                                    else:
                                        sh.cell(row=2,column=updatecol[checky]).fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                                for checky2 in range(len(updatefill)):
                                    checkcolumnsrc=sh.cell(row=2,column=updatefill[checky2]).value
                                    if checkcolumnsrc in Colnames2:
                                        if found==1:
                                            found=1
                                        else:
                                            found=1
                                        if sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).value is None:
                                               sh2.cell(row=s,column=Colnames2[columnkey]+1).fill=PatternFill(start_color="CEFFC7", end_color="CEFFC7", fill_type = "solid")
                                               sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).value=sh.cell(row=r,column=updatefill[checky2]).value
                                               sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).fill=PatternFill(start_color="CEFFC7", end_color="CEFFC7", fill_type = "solid")
                                        if isinstance(sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).value, datetime.datetime) is True:
                                            sh2.cell(row=s,column=Colnames2[checkcolumnsrc]+1).number_format='yyyy-mm-dd'
                                    else:
                                        sh.cell(row=2,column=updatefill[checky2]).fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                            s+=1
                    else:
                        wrkbk2.close()
                        break
                    wrkbk2.save(g)
                    wrkbk2.close()
                    if found==1:
                        sh.cell(row=r,column=Keys['key']+1).fill=PatternFill(start_color="CEFFC7", end_color="CEFFC7", fill_type = "solid")
                        countfound=countfound+1
                    else:
                        sh.cell(row=r,column=Keys['key']+1).fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                    iterations=iterations+1
                    print("Current Progress "+str(round(((iterations/maxprogress)*100),2))+"%")
                r+=1
            wrkbk.save(f)
            wrkbk.close()
            print()
            print("Done,Processed "+str(countfound)+" Rows out of "+str(maxprogress))
            print("Kindly Close This Window")
    except:
        print("Check the file and try again")
# open button
open_button = ttk.Button(
    root,
    text='Select Source',
    command=select_source
)
target_button = ttk.Button(
    root,
    text='Select Target',
    command=select_target
)
run_button = ttk.Button(
    root,
    text='Run Program',
    command=run_program
)

open_button.place(x=20,y=60)
target_button.place(x=20,y=120)
run_button.place(x=20,y=180)
root.mainloop()
